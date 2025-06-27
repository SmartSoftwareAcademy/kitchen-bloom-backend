from typing import Dict, List, Any, Optional, Tuple
import os
import tempfile
import shutil
import polars as pl
from decimal import Decimal
from datetime import datetime
from django.conf import settings
from django.template.loader import render_to_string
from weasyprint import HTML, CSS
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image
from openpyxl.chart import LineChart, Reference
from openpyxl.chart.layout import Layout, ManualLayout
from apps.branches.models import Company, Branch
from apps.base.utils import get_request_branch_id
import csv


class ReportExporter:
    """Class for exporting reports in different formats."""
    
    def export_to_pdf(self, report_type: str, report_data: Dict[str, Any],
                     template_name: str, filename: str, request=None) -> str:
        """
        Export report data to PDF format.
        
        Args:
            report_type: Type of the report
            report_data: Data to include in the report
            template_name: Name of the template to use
            filename: Name for the output file
            request: Django request object
            
        Returns:
            Path to the generated PDF file
        """
        # Fetch company/branch info
        branch = None
        company = None
        branch_id = None
        if request:
            branch_id = get_request_branch_id(request)
        if branch_id:
            try:
                branch = Branch.objects.select_related('company').get(id=branch_id)
                company = branch.company
            except Branch.DoesNotExist:
                company = Company.objects.filter(is_active=True).first()
        else:
            company = Company.objects.filter(is_active=True).first()
            branch = company.branches.filter(is_default=True).first() if company else None
        
        # Render HTML template
        html_content = render_to_string(
            f'reporting/reports/{template_name}.html',
            {
                'report_data': report_data,
                'company': company,
                'branch': branch,
                'company_logo_url': company.logo.url if company and company.logo else None
            }
        )
        
        # Create temporary directory
        temp_dir = tempfile.mkdtemp()
        
        try:
            # Generate PDF
            pdf_file = os.path.join(temp_dir, f'{filename}.pdf')
            HTML(string=html_content).write_pdf(
                pdf_file,
                stylesheets=[
                    CSS(string='''
                        @page {
                            size: A4;
                            margin: 2cm;
                        }
                        ''')
                ]
            )
            
            # Copy to final destination
            final_path = os.path.join(settings.MEDIA_ROOT, 'reports', f'{filename}.pdf')
            os.makedirs(os.path.dirname(final_path), exist_ok=True)
            shutil.copy2(pdf_file, final_path)
            
            return final_path
            
        finally:
            # Clean up temporary files
            shutil.rmtree(temp_dir)
    
    def export_to_excel(self, report_type: str, report_data: Dict[str, Any],
                       filename: str) -> str:
        """
        Export report data to Excel format.
        
        Args:
            report_type: Type of the report
            report_data: Data to include in the report
            filename: Name for the output file
            
        Returns:
            Path to the generated Excel file
        """
        # Create workbook and worksheet
        wb = Workbook()
        ws = wb.active
        ws.title = report_type
        
        # Add header row
        headers = [
            'Period', 'Total Sales', 'Order Count', 'Average Order'
        ]
        
        # Style header
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.value = header
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")
        
        # Add data rows
        for row_num, period in enumerate(report_data['time_series'], 2):
            ws.cell(row=row_num, column=1, value=period['period'])
            ws.cell(row=row_num, column=2, value=float(period['total_sales']))
            ws.cell(row=row_num, column=3, value=period['order_count'])
            ws.cell(row=row_num, column=4, value=float(period['avg_order']))
        
        # Add totals row
        totals_row = len(report_data['time_series']) + 2
        ws.cell(row=totals_row, column=1, value='Total')
        ws.cell(row=totals_row, column=2, value=float(report_data['totals']['total_sales']))
        ws.cell(row=totals_row, column=3, value=report_data['totals']['order_count'])
        ws.cell(row=totals_row, column=4, value=float(report_data['totals']['avg_order']))
        
        # Add chart
        chart = LineChart()
        chart.title = "Sales Trend"
        chart.style = 13
        chart.x_axis.title = "Period"
        chart.y_axis.title = "Amount ($)"
        
        # Add data to chart
        data = Reference(ws, min_col=2, min_row=1, max_col=2, max_row=len(report_data['time_series']) + 1)
        chart.add_data(data, titles_from_data=True)
        
        # Add categories
        cats = Reference(ws, min_col=1, min_row=2, max_row=len(report_data['time_series']) + 1)
        chart.set_categories(cats)
        
        # Add chart to worksheet
        ws.add_chart(chart, "F2")
        
        # Adjust column widths
        for col_num in range(1, len(headers) + 1):
            col_letter = get_column_letter(col_num)
            ws.column_dimensions[col_letter].width = 15
        
        # Save workbook
        temp_dir = tempfile.mkdtemp()
        try:
            excel_file = os.path.join(temp_dir, f'{filename}.xlsx')
            wb.save(excel_file)
            
            # Copy to final destination
            final_path = os.path.join(settings.MEDIA_ROOT, 'reports', f'{filename}.xlsx')
            os.makedirs(os.path.dirname(final_path), exist_ok=True)
            shutil.copy2(excel_file, final_path)
            
            return final_path
            
        finally:
            # Clean up temporary files
            shutil.rmtree(temp_dir)
    
    def export_to_csv(self, report_type: str, report_data: Dict[str, Any],
                     filename: str) -> str:
        """
        Export report data to CSV format.
        
        Args:
            report_type: Type of the report
            report_data: Data to include in the report
            filename: Name for the output file
            
        Returns:
            Path to the generated CSV file
        """
        # Create Polars DataFrame
        df = pl.DataFrame(report_data['time_series'])
        df = df.with_column(
            pl.col('period').cast(pl.Date).dt.strftime('%Y-%m-%d')
        )
        
        # Add totals row as a new DataFrame and concatenate
        totals_df = pl.DataFrame({
            'period': ['Total'],
            'total_sales': [report_data['totals']['total_sales']],
            'order_count': [report_data['totals']['order_count']],
            'avg_order': [report_data['totals']['avg_order']]
        })
        df = pl.concat([df, totals_df])
        
        # Create temporary directory
        temp_dir = tempfile.mkdtemp()
        try:
            # Generate CSV
            csv_file = os.path.join(temp_dir, f'{filename}.csv')
            df.write_csv(csv_file)
            
            # Copy to final destination
            final_path = os.path.join(settings.MEDIA_ROOT, 'reports', f'{filename}.csv')
            os.makedirs(os.path.dirname(final_path), exist_ok=True)
            shutil.copy2(csv_file, final_path)
            
            return final_path
            
        finally:
            # Clean up temporary files
            shutil.rmtree(temp_dir)

    def export_dashboard_to_csv(self, report_data: Dict[str, Any], filename: str) -> str:
        """
        Export dashboard data to CSV format.
        
        Args:
            report_data: Dashboard data structure
            filename: Name for the output file
            
        Returns:
            Path to the generated CSV file
        """
        # Create temporary directory
        temp_dir = tempfile.mkdtemp()
        try:
            # Generate CSV
            csv_file = os.path.join(temp_dir, f'{filename}.csv')
            
            with open(csv_file, 'w', newline='', encoding='utf-8') as f:
                writer = csv.writer(f)
                
                # Write header
                writer.writerow(['Dashboard Overview Report'])
                writer.writerow([])
                
                # Sales Summary
                if 'sales_summary' in report_data:
                    writer.writerow(['Sales Summary'])
                    sales_summary = report_data['sales_summary']
                    if 'time_series' in sales_summary:
                        writer.writerow(['Period', 'Total Sales', 'Order Count', 'Average Order'])
                        for period in sales_summary['time_series']:
                            writer.writerow([
                                period.get('period', ''),
                                period.get('total_sales', 0),
                                period.get('order_count', 0),
                                period.get('avg_order', 0)
                            ])
                        if 'totals' in sales_summary:
                            totals = sales_summary['totals']
                            writer.writerow(['Total', totals.get('total_sales', 0), totals.get('order_count', 0), totals.get('avg_order', 0)])
                    writer.writerow([])
                
                # Top Products
                if 'top_products' in report_data:
                    writer.writerow(['Top Products'])
                    writer.writerow(['Product', 'Quantity Sold', 'Revenue', 'Percentage'])
                    for product in report_data['top_products']:
                        writer.writerow([
                            product.get('product_name', ''),
                            product.get('quantity_sold', 0),
                            product.get('revenue', 0),
                            f"{product.get('percentage', 0)}%"
                        ])
                    writer.writerow([])
                
                # Stock Alerts
                if 'stock_alerts' in report_data:
                    writer.writerow(['Stock Alerts'])
                    writer.writerow(['Product', 'Current Stock', 'Threshold', 'Status'])
                    for alert in report_data['stock_alerts']:
                        writer.writerow([
                            alert.get('product_name', ''),
                            alert.get('current_stock', 0),
                            alert.get('threshold', 0),
                            alert.get('status', '')
                        ])
                    writer.writerow([])
                
                # Financial Summary
                if 'profit_data' in report_data:
                    writer.writerow(['Profit Summary'])
                    profit_data = report_data['profit_data']
                    if 'summary' in profit_data:
                        summary = profit_data['summary']
                        writer.writerow(['Metric', 'Value'])
                        writer.writerow(['Total Revenue', summary.get('total_revenue', 0)])
                        writer.writerow(['Total Expenses', summary.get('total_expenses', 0)])
                        writer.writerow(['Net Profit', summary.get('net_profit', 0)])
                        writer.writerow(['Profit Margin', f"{summary.get('profit_margin', 0)}%"])
                    writer.writerow([])
                
                if 'expense_data' in report_data:
                    writer.writerow(['Expense Summary'])
                    expense_data = report_data['expense_data']
                    if 'summary' in expense_data:
                        summary = expense_data['summary']
                        writer.writerow(['Metric', 'Value'])
                        writer.writerow(['Total Expenses', summary.get('total_expenses', 0)])
                        writer.writerow(['Average Daily Expense', summary.get('avg_daily_expense', 0)])
            
            # Copy to final destination
            final_path = os.path.join(settings.MEDIA_ROOT, 'reports', f'{filename}.csv')
            os.makedirs(os.path.dirname(final_path), exist_ok=True)
            shutil.copy2(csv_file, final_path)
            
            return final_path
            
        finally:
            # Clean up temporary files
            shutil.rmtree(temp_dir)
