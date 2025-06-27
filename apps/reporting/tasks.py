import logging
import os
import tempfile
from datetime import datetime
from typing import Any, Dict, Optional

from celery import shared_task
from django.conf import settings
from django.core.files.base import ContentFile
from django.core.files.storage import default_storage
from django.template.loader import render_to_string
from django.utils import timezone
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, Side
from openpyxl.utils import get_column_letter
import polars as pl
from weasyprint import HTML

from .models import Report, ReportExecutionLog, ReportSchedule
from .utils import get_report_data

logger = logging.getLogger(__name__)


@shared_task(bind=True, max_retries=3, default_retry_delay=300)
def generate_report_async(
    self,
    report_id: int,
    user_id: int,
    execution_log_id: int,
    schedule_id: Optional[int] = None,
    **kwargs
) -> Dict[str, Any]:
    """
    Asynchronous task to generate a report.
    
    Args:
        report_id: ID of the report to generate
        user_id: ID of the user who requested the report
        execution_log_id: ID of the execution log for this report generation
        schedule_id: Optional ID of the schedule that triggered this report
        **kwargs: Additional arguments for the report
    
    Returns:
        Dict with the status and file path of the generated report
    """
    from django.contrib.auth import get_user_model
    User = get_user_model()
    
    try:
        # Get the report and execution log
        report = Report.objects.get(id=report_id)
        execution_log = ReportExecutionLog.objects.get(id=execution_log_id)
        user = User.objects.get(id=user_id)
        
        # Update execution log
        execution_log.status = 'running'
        execution_log.started_at = timezone.now()
        execution_log.save()
        
        # Get report data using the utility function
        report_data = get_report_data(report, user)
        
        # Generate the report file
        file_path, file_size = _generate_report_file(report, report_data, user)
        
        # Update execution log with success
        execution_log.status = 'completed'
        execution_log.completed_at = timezone.now()
        execution_log.file_path = file_path
        execution_log.file_size = file_size
        execution_log.save()
        
        # If this was triggered by a schedule, update the next run time
        if schedule_id:
            try:
                schedule = ReportSchedule.objects.get(id=schedule_id)
                schedule.update_next_run()
            except ReportSchedule.DoesNotExist:
                pass
        
        return {
            'status': 'success',
            'file_path': file_path,
            'file_size': file_size
        }
        
    except Exception as e:
        logger.error(f"Error generating report: {str(e)}", exc_info=True)
        
        # Update execution log with failure
        try:
            execution_log.status = 'failed'
            execution_log.error_message = str(e)
            execution_log.completed_at = timezone.now()
            execution_log.save()
        except Exception as update_error:
            logger.error(f"Failed to update execution log: {str(update_error)}")
            
        # Retry the task if we haven't exceeded max retries
        if self.request.retries < self.max_retries:
            raise self.retry(exc=e, countdown=60 * (self.request.retries + 1))
            
        return {
            'status': 'error',
            'error': str(e)
        }


def _generate_report_file(
    report: Report,
    report_data: Dict[str, Any],
    user: Any
) -> tuple[str, int]:
    """
    Generate a report file in the requested format.
    
    Args:
        report: The report to generate
        report_data: The data to include in the report
        user: The user who requested the report
    
    Returns:
        Tuple of (file_path, file_size)
    """
    # Determine the output format (default to PDF)
    output_format = getattr(report, 'format', 'pdf').lower()
    
    # Create a unique filename
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    filename = f"{report.name.lower().replace(' ', '_')}_{timestamp}"
    
    # Generate the file based on the requested format
    if output_format == 'pdf':
        return _generate_pdf(report, report_data, user, filename)
    elif output_format == 'excel':
        return _generate_excel(report, report_data, user, filename)
    elif output_format == 'csv':
        return _generate_csv(report, report_data, user, filename)
    elif output_format == 'json':
        return _generate_json(report, report_data, user, filename)
    else:
        # Default to PDF if format is not recognized
        return _generate_pdf(report, report_data, user, filename)


def _generate_pdf(
    report: Report,
    report_data: Dict[str, Any],
    user: Any,
    filename: str
) -> tuple[str, int]:
    """Generate a PDF report."""
    # Render the HTML template
    context = {
        'report': report,
        'data': report_data,
        'user': user,
        'generated_at': timezone.now(),
        'title': report.name,
    }
    
    # Render the template
    html_string = render_to_string('reporting/pdf_template.html', context)
    
    # Create a temporary file
    with tempfile.NamedTemporaryFile(delete=False, suffix='.pdf') as temp_file:
        # Generate PDF
        HTML(string=html_string, base_url=settings.BASE_DIR).write_pdf(temp_file)
        temp_file_path = temp_file.name
    
    # Save to storage
    file_path = f'reports/pdf/{filename}.pdf'
    with open(temp_file_path, 'rb') as f:
        file = default_storage.save(file_path, ContentFile(f.read()))
    
    # Clean up the temporary file
    try:
        os.unlink(temp_file_path)
    except OSError:
        pass
    
    # Get file size
    file_size = default_storage.size(file_path)
    
    return file_path, file_size


def _generate_excel(
    report: Report,
    report_data: Dict[str, Any],
    user: Any,
    filename: str
) -> tuple[str, int]:
    """
    Generate an Excel report using polars.
    """
    try:
        # Create a new workbook and select the active worksheet
        wb = Workbook()
        ws = wb.active
        ws.title = report.name[:31]  # Excel sheet name max length is 31

        # Add report header
        ws.append([f"Report: {report.name}"])
        ws.append([f"Generated on: {timezone.now().strftime('%Y-%m-%d %H:%M:%S')}"])
        ws.append([f"Generated by: {user.get_full_name()}"])
        ws.append([""])  # Empty row

        # Convert report data to a DataFrame using polars
        if 'data' in report_data and isinstance(report_data['data'], (list, dict)):
            # Convert to polars DataFrame
            if isinstance(report_data['data'], dict):
                df = pl.DataFrame(report_data['data'])
            else:  # list of dicts
                df = pl.from_dicts(report_data['data'])
            
            # Write column headers
            ws.append(df.columns)
            
            # Write data rows
            for row in df.rows():
                ws.append(row)
            
            # Apply formatting to header row
            for col in range(1, len(df.columns) + 1):
                cell = ws.cell(row=5, column=col)  # Header is on row 5 (after 4 header rows)
                cell.font = Font(bold=True)
                cell.border = Border(
                    left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin')
                )
                cell.alignment = Alignment(horizontal='center')
                
                # Auto-size columns
                column_letter = get_column_letter(col)
                max_length = max(
                    df[df.columns[col-1]].cast(pl.Utf8).str.lengths().max(),
                    len(str(df.columns[col-1]))
                )
                adjusted_width = (max_length + 2) * 1.2
                ws.column_dimensions[column_letter].width = adjusted_width
        else:
            # If data is not in the expected format, just write it as is
            ws.append(["Report Data:"])
            for key, value in report_data.items():
                ws.append([f"{key}: {value}"])

        # Save the workbook to a temporary file
        with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as tmp:
            temp_path = tmp.name
            wb.save(temp_path)
            
        # Read the file content
        with open(temp_path, 'rb') as f:
            content = f.read()
            
        # Clean up
        os.unlink(temp_path)
        
        # Save to storage
        file_path = default_storage.save(
            f'reports/{filename}.xlsx',
            ContentFile(content)
        )
        
        return file_path, len(content)
        
    except Exception as e:
        logger.error(f"Error generating Excel report: {str(e)}", exc_info=True)
        raise
    
    # Get file size
    file_size = default_storage.size(file_path)
    
    return file_path, file_size


def _generate_csv(
    report: Report,
    report_data: Dict[str, Any],
    user: Any,
    filename: str
) -> tuple[str, int]:
    """
    Generate a CSV report using polars.
    """
    try:
        # Convert report data to a polars DataFrame
        if 'data' in report_data and isinstance(report_data['data'], (list, dict)):
            if isinstance(report_data['data'], dict):
                df = pl.DataFrame(report_data['data'])
            else:  # list of dicts
                df = pl.from_dicts(report_data['data'])
            
            # Create CSV content
            csv_content = df.write_csv()
        else:
            # If data is not in the expected format, create a simple CSV
            csv_content = "key,value\n"
            for key, value in report_data.items():
                csv_content += f"{key},{value}\n"
        
        # Save to storage
        file_path = default_storage.save(
            f'reports/{filename}.csv',
            ContentFile(csv_content.encode('utf-8'))
        )
        
        return file_path, len(csv_content)
        
    except Exception as e:
        logger.error(f"Error generating CSV report: {str(e)}", exc_info=True)
        raise


def _generate_json(
    report: Report,
    report_data: Dict[str, Any],
    user: Any,
    filename: str
) -> tuple[str, int]:
    """Generate a JSON report."""
    import json
    
    # Prepare the data
    output_data = {
        'report': {
            'id': report.id,
            'name': report.name,
            'type': report.report_type,
        },
        'generated_by': user.get_full_name() or user.username,
        'generated_at': timezone.now().isoformat(),
        'data': report_data
    }
    
    # Convert to JSON
    json_data = json.dumps(output_data, indent=2, default=str)
    
    # Save to storage
    file_path = default_storage.save(
        f'reports/{filename}.json',
        ContentFile(json_data.encode('utf-8'))
    )
    
    # Get file size
    file_size = default_storage.size(file_path)
    
    return file_path, file_size
